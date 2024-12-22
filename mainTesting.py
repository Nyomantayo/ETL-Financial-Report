import tkinter as tk
import mysql.connector
import pandas as pd
import pdfplumber
import openpyxl
import fitz
import re
from pathlib import Path
from tkinter import filedialog, messagebox, Listbox

# ======= Global Variabel ======
grup_laporan_keuangan = ["Laporan Neraca", "Laporan Laba Rugi", "Laporan Arus Kas"] # Informasi grup laporan keuangan
word_check = "CATATAN ATAS LAPORAN KEUANGAN"
word_boundaries = "lain)"
kode_emiten_global = None

# Excel
excel_path = None
data_neraca = None
data_aruskas = None
data_labarugi = None
status_neraca = None
status_aruskas = None
status_labarugi = None
status_kuartal_unik = None
informasi_sheet_keuangan = [] # Berisi Sheet yang menunjukkan Laporan Posisi Keuangan(Neraca), Laba Rugi, dan Arus Kas
selected_files_excel = [] # Menyimpan daftar file Excel yang dipilih
informasi_tambahan = [] # Informasi kode emiten, nama emite, quartal, dan tahun

# PDF
selected_files_pdf = []   # Menyimpan daftar file PDF yang dipilih
teks_labarugi = []
teks_aruskas = []
teks_neraca = []
data = []

# ======= Bagian DATABASE ==========
def masukkan_ke_database():
    global status_kuartal_unik
    try:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='db_keuangan'
        )

        cursor = conn.cursor()
        sql = "INSERT INTO tb_laporan_keuangan (kode_emiten, nama_emiten, tahun, quartal, grup_laporan_keuangan, item, nilai) VALUES (%s, %s, %s, %s, %s, %s, %s)"

        # Laporan Neraca
        validasi_data_keuangan(grup_laporan_keuangan[0])
        if status_neraca == False:
            for i in range(3, len(data_neraca)):
                nilai_jutaan = data_neraca.iloc[i, 1]
                if status_kuartal_unik == False:
                    nilai_jutaan = data_neraca.iloc[i, 1] * 1000000
                cursor.execute(sql, (informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], grup_laporan_keuangan[0], data_neraca.iloc[i,0], nilai_jutaan))
                conn.commit()
        
        # Laporan Laba Rugi
        validasi_data_keuangan(grup_laporan_keuangan[1])
        if status_labarugi == False:
            for i in range(3, len(data_labarugi)):
                nilai_jutaan = data_labarugi.iloc[i, 1]
                if status_kuartal_unik == False:
                    nilai_jutaan = data_labarugi.iloc[i, 1] * 1000000 
                cursor.execute(sql, (informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], grup_laporan_keuangan[1], data_labarugi.iloc[i,0], nilai_jutaan))
                conn.commit()
        
        # Laporan Arus Kas
        validasi_data_keuangan(grup_laporan_keuangan[2])
        if status_aruskas == False:
            for i in range(3, len(data_aruskas)):
                nilai_jutaan = data_aruskas.iloc[i, 1]
                if status_kuartal_unik == False:
                    nilai_jutaan = data_aruskas.iloc[i, 1] * 1000000
                cursor.execute(sql, (informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], grup_laporan_keuangan[2], data_aruskas.iloc[i,0], nilai_jutaan))
                conn.commit()
        
        print(cursor.rowcount, "Data berhasil disimpan")

        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Error", f"Gagal Memasukkan ke Database: {e}")

def cocokkan_database_pdf(teks_keuangan):
    try:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='db_keuangan'
        )

        cursor = conn.cursor()

        for index, isi in enumerate(teks_keuangan):
            kata_kunci = f"%{isi}%"
            # print(f"index-{index}: {isi}")
            sql = """
                SELECT * 
                FROM tb_laporan_keuangan 
                WHERE 
                LOWER(nama_emiten) = LOWER(%s) AND 
                tahun = %s AND 
                quartal = %s AND 
                LOWER(grup_laporan_keuangan) = LOWER(%s) AND 
                LOWER(item) LIKE LOWER(%s)
            """

            # Eksekusi query SELECT
            cursor.execute(sql, (informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], kata_kunci))
            hasil_select = cursor.fetchall()  # Mengambil satu hasil query
            
            if hasil_select:
                # Mencari Hasil yang telah didapat (mencari nama item)
                # print("Masuk Hasil Select")
                kolom_nama = [desc[0] for desc in cursor.description]
                baris_satu = hasil_select[0]  # Baris ke-1
                kolom_nama_item = dict(zip(kolom_nama, baris_satu))['item']
                
                if index+1 <= len(teks_keuangan) and cek_pola(teks_keuangan[index+1]):
                    print(f"[1] Ini Bisa dimasukkan! {isi} {teks_keuangan[index+1]}")
                    try:
                        temp = teks_keuangan[index+1].split(',')
                        for index, isi in enumerate(temp):
                            if index == 0:
                                sql = """
                                    UPDATE tb_laporan_keuangan 
                                    SET notes = %s
                                    WHERE 
                                    LOWER(nama_emiten) = LOWER(%s) AND 
                                    tahun = %s AND 
                                    quartal = %s AND 
                                    LOWER(grup_laporan_keuangan) = LOWER(%s) AND 
                                    LOWER(item) LIKE LOWER(%s)
                                """

                                # Eksekusi query UPDATE
                                cursor.execute(sql, (isi, informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], kata_kunci))
                                conn.commit()
                            else:
                                sql = "INSERT INTO tb_laporan_keuangan (kode_emiten, nama_emiten, tahun, quartal, grup_laporan_keuangan, item, nilai, notes) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                                # Eksekusi query UPDATE
                                cursor.execute(sql, ("BBNI", informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], kolom_nama_item, 0, isi))
                                conn.commit()
                    except Exception as e:
                        messagebox.showerror("Error", f"Gagal Mengupdate Database: {e}")
                elif index+2 <= len(teks_keuangan) and cek_pola(teks_keuangan[index+2]):
                    print(f"[2] Ini Bisa dimasukkan! {isi} {teks_keuangan[index+2]}")
                    try:
                        temp = teks_keuangan[index+2].split(',')
                        for index, isi in enumerate(temp):
                            if index == 0:
                                sql = """
                                    UPDATE tb_laporan_keuangan 
                                    SET notes = %s
                                    WHERE 
                                    LOWER(nama_emiten) = LOWER(%s) AND 
                                    tahun = %s AND 
                                    quartal = %s AND 
                                    LOWER(grup_laporan_keuangan) = LOWER(%s) AND 
                                    LOWER(item) LIKE LOWER(%s)
                                """

                                # Eksekusi query UPDATE
                                cursor.execute(sql, (isi, informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], kata_kunci))
                                conn.commit()
                            else:
                                sql = "INSERT INTO tb_laporan_keuangan (kode_emiten, nama_emiten, tahun, quartal, grup_laporan_keuangan, item, nilai, notes) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                                # Eksekusi query UPDATE
                                cursor.execute(sql, ("BBNI", informasi_tambahan[0], informasi_tambahan[1], informasi_tambahan[2], informasi_tambahan[3], kolom_nama_item, 0, isi))
                                conn.commit()
                    except Exception as e:
                        messagebox.showerror("Error", f"Gagal Mengupdate Database: {e}")
        messagebox.showinfo("Berhasil", "Data Berhasil Diupdate: ")
    except Exception as e:
        messagebox.showerror("Error", f"Gagal Mencocokkan Database: {e}")
    
    cursor.close()
    conn.close()

def hilangkan_nan(df):
    df.fillna("0", inplace=True)
    return df

def hilangkan_nan_strip(df):
    df.fillna("-", inplace=True)
    return df

def cek_pola(teks):
    # Regex untuk mencocokkan teks dengan pola 1a hingga 100z
    pola = r"^(\d{1,2}|(\d{1,2}[a-z])|(\d{1,2}(?:,\d{1,2}[a-z]?)*))$"

    # Memeriksa apakah teks cocok dengan pola
    if re.match(pola, teks):
        return True
    else:
        return False

# ======= Bagian PDF ==========
def pilih_file_pdf():
    # Jika masih ada informasi tersisa hapus
    informasi_tambahan.clear()
    informasi_sheet_keuangan.clear()

    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    for file_path in file_paths:
        if file_path not in selected_files_pdf:
            selected_files_pdf.append(file_path)
            file_listbox.insert(tk.END, f"PDF: {file_path}")

def cari_halaman_mengandung_teks(nama_file, kata_kunci):
    kata_kunci = str(kata_kunci).lower()
    halaman_ditemukan = []
    doc = fitz.open(nama_file)
    
    for halaman_num in range(doc.page_count):
        halaman = doc[halaman_num]
        teks = halaman.get_text()
        teks = str(teks).lower()
        if kata_kunci in teks:
            halaman_ditemukan.append(halaman_num + 1)  # fitz menggunakan indeks 0, jadi tambahkan 1 untuk halaman yang benar

    doc.close()

    # Hapus halaman yang non berurutan
    i = 1
    while i < len(halaman_ditemukan):
        if halaman_ditemukan[i] - halaman_ditemukan[i - 1] > 1:
            del halaman_ditemukan[i]
        else:
            i += 1
    return halaman_ditemukan

def cari_halaman_mengandung_teks_pdf(nama_file, kata_kunci):
    kata_kunci = str(kata_kunci).lower()
    halaman_ditemukan = []
    doc = fitz.open(nama_file)
    
    for halaman_num in range(doc.page_count):
        halaman = doc[halaman_num]
        teks = halaman.get_text()
        teks = str(teks).lower()
        if kata_kunci in teks:
            halaman_ditemukan.append(halaman_num + 1)  # fitz menggunakan indeks 0, jadi tambahkan 1 untuk halaman yang benar

    doc.close()

    return halaman_ditemukan

def ambil_teks_di_halaman(nama_file, page_number):
    try:
        # Membuka file PDF
        doc = fitz.open(nama_file)

        # Cek Apakah Nomor Halaman Ada
        if page_number < 1 or page_number > len(doc):
            print(f"Halaman {page_number} tidak valid. PDF ini memiliki {len(doc)} halaman.")
            return
        
        # Mengakses halaman tertentu (zero-based index)
        page = doc[page_number - 1]

        # Mengekstrak teks
        text = page.get_text()

        # Filter teks untuk menghilangkan, baris kosong, dan spasi berlebih
        filtered_text = []
        for line in text.splitlines():
            # Menghilangkan baris kosong
            if not line.strip():
                continue
            
            # Menghilangkan spasi berlebih
            cleaned_line = re.sub(r'\s+', ' ', line.strip())

            # Menyimpan di variabel filtered text
            filtered_text.append(cleaned_line)

        return filtered_text
    except Exception as e:
        print(f"Terjadi kesalahan: {e}")

def cari_katakunci_di_teks(kata_kunci, teks):
    kata_kunci = str(kata_kunci).lower()
    for isi in teks:
        temp = str(isi).lower()
        if kata_kunci in temp:
            return True
    return False

def proses_files_pdf():
    global teks_neraca, teks_labarugi, teks_aruskas
    teks_informasi = []
    for file_path in selected_files_pdf:
        # Membersihkan Hal-Hal yang Tersisa
        informasi_tambahan.clear()
        informasi_sheet_keuangan.clear()
        teks_neraca.clear()
        teks_labarugi.clear()
        teks_aruskas.clear()

        try:
            # Cari Informasi Umum dari PDF
            # Informasi Tambahan : Nama Emiten, Tahun, Quartal, Grup Keuangan
            for i in range(1, 20):
                teks_informasi.extend(ambil_teks_di_halaman(file_path, i))

            # Nama Emiten ada di baris ke 0 halaman 1
            if teks_informasi:
                informasi_tambahan.append(teks_informasi[0])
            else:
                informasi_tambahan.append("PT BANK NEGARA INDONESIA (PERSERO) Tbk")

            # Mencari Tahun di halaman 1
            for i in range(2026, 2019, -1):
                hasil = cari_katakunci_di_teks(i, teks_informasi)
                if hasil:
                    informasi_tambahan.append(i)
                    break
            
            # Mencari Quartal di halaman 1
            if cari_katakunci_di_teks("Maret", teks_informasi):
                informasi_tambahan.append(1)
            elif cari_katakunci_di_teks("Juni", teks_informasi):
                informasi_tambahan.append(2)
            elif cari_katakunci_di_teks("September", teks_informasi):
                informasi_tambahan.append(3)
            else:
                informasi_tambahan.append(4)
            
            # Mencari dan Mengisi Informasi Grup Keuangan
            # Laporan Neraca
            halaman_neraca = cari_halaman_mengandung_teks_pdf(file_path, "Laporan Posisi Keuangan")
            for i in halaman_neraca:
                teks_neraca.extend(ambil_teks_di_halaman(file_path, i))
            informasi_tambahan.append("Laporan Neraca")
            print(informasi_tambahan)
            cocokkan_database_pdf(teks_neraca)
            
            # Laporan Laba Rugi
            halaman_labarugi = cari_halaman_mengandung_teks_pdf(file_path, "Laporan Laba Rugi")
            for i in halaman_labarugi:
                teks_labarugi.extend(ambil_teks_di_halaman(file_path, i))
            informasi_tambahan[3] = "Laporan Laba Rugi"
            print(informasi_tambahan)
            cocokkan_database_pdf(teks_labarugi)

            # Laporan Arus Kas
            halaman_aruskas = cari_halaman_mengandung_teks_pdf(file_path, "Laporan Arus Kas")
            for i in halaman_aruskas:
                teks_aruskas.extend(ambil_teks_di_halaman(file_path, i))
            informasi_tambahan[3] = "Laporan Arus Kas"
            print(informasi_tambahan)
            cocokkan_database_pdf(teks_aruskas)

        except Exception as e:
            messagebox.showerror("Error", f"Gagal membaca file PDF: {e}")

def ambil_kode_emiten_nama_file(nama_file):
    match = re.search(r'[A-Z]{4}', nama_file)
    if match:
        return match.group(0)  # Mengembalikan 4 huruf kapital pertama yang ditemukan
    return None

# ======= Bagian EXCEL ==========
def pilih_file_excel():
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx;*.xls")])
    for file_path in file_paths:
        if file_path not in selected_files_excel:
            selected_files_excel.append(file_path)
            file_listbox.insert(tk.END, f"Excel: {file_path}")

def cari_sheet_excel(nama_file_excel, kalimat_dicari):
    kalimat_dicari = str(kalimat_dicari).lower()
    workbook = openpyxl.load_workbook(nama_file_excel)
    for sheet_index, sheet in enumerate(workbook.sheetnames):
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if kalimat_dicari in cell_value:
                        return sheet

def proses_files_excel():
    global status_aruskas, status_neraca, status_labarugi
    global excel_path, status_kuartal_unik
    # Untuk mencari informasi umum
    for file_path in selected_files_excel:
        try:
            # Set Semua Status sebagai False untuk menerima data
            status_neraca = False
            status_aruskas = False
            status_labarugi = False
            status_kuartal_unik = False

            # Membersihkan Hal-Hal yang Tersisa
            informasi_tambahan.clear()
            informasi_sheet_keuangan.clear()
            teks_neraca.clear()
            teks_labarugi.clear()
            teks_aruskas.clear()

            # Mencari Sheet Informasi Umum dengan kata kunci "Kode Entitas"
            sheet_informasi_umum = cari_sheet_excel(file_path, "Kode Entitas")
            df = pd.read_excel(file_path, sheet_name = sheet_informasi_umum)
            for i in range(1, len(df)):
                if(df.iloc[i,0] == "Kode entitas"):
                    kode_emiten = df.iloc[i,1]
                if(df.iloc[i,0] == "Nama entitas"):
                    nama_emiten = df.iloc[i,1]
                    if not nama_emiten[:2].upper() == "PT":
                        nama_emiten = "PT " + nama_emiten
                if(df.iloc[i,0] == "Periode penyampaian laporan keuangan"):
                    if(df.iloc[i,1] == "Kuartal I / First Quarter"):
                        quartal = 1
                    elif (df.iloc[i,1] == "Kuartal II / Second Quarter"):
                        quartal = 2
                    elif (df.iloc[i,1] == "Kuartal III / Third Quarter"):
                        quartal = 3
                    else:
                        quartal = 4
                        status_kuartal_unik = True
                if(df.iloc[i,0] == "Tanggal awal periode berjalan"):
                    tahun = int(str(df.iloc[i,1])[:4])
            
            # Menyimpan data ke informasi tambahan
            informasi_tambahan.append(kode_emiten)
            informasi_tambahan.append(nama_emiten)
            informasi_tambahan.append(tahun)
            informasi_tambahan.append(quartal)

            # Mencari Sheet Grup Laporan Keuangan
            kata_kunci = ["LAPORAN POSISI KEUANGAN", "LAPORAN LABA RUGI", "LAPORAN ARUS KAS"]
            for kata in kata_kunci:
                sheet_keuangan = cari_sheet_excel(file_path, kata)
                informasi_sheet_keuangan.append(sheet_keuangan)
            
            # Set Excel Path Saat ini
            excel_path = file_path

            # Mengambil Data Keperluan (Neraca, Laba Rugi, Arus Kas)
            ambil_keperluan_excel()

            # Memasukkan ke database
            masukkan_ke_database()
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membaca file Excel: {e}")

def baca_file_excel(nama_file, sheet_name):
    df = pd.read_excel(nama_file, sheet_name=sheet_name)
    return df

def ambil_keperluan_excel():
    global data_neraca, data_aruskas, data_labarugi
    for idx, isi in enumerate(informasi_sheet_keuangan):
        if idx == 0:
            data_neraca = baca_file_excel(excel_path, isi)
            data_neraca = hilangkan_nan(data_neraca)
        elif idx == 1:
            data_labarugi = baca_file_excel(excel_path, isi)
            data_labarugi = hilangkan_nan(data_labarugi)
        else: 
            data_aruskas = baca_file_excel(excel_path, isi)
            data_aruskas = hilangkan_nan(data_aruskas)

# ======= Bagian CALK ==========
def search_pages(pdf_path, word_check):
    pages_with_text = []

    plumber_reader = pdfplumber.open(pdf_path)
    pdf_document = fitz.open(pdf_path)
    total_pages = pdf_document.page_count
    pages = range(total_pages)

    boundaries = 0
    print("[INFO] MELAKUKAN PENGECEKAN HALAMAN")
    for page_num in pages:
        print(f"Mengecek Halaman {page_num + 1}...")
        page = pdf_document[page_num]
        page_text = page.get_text("text")

        if word_check in page_text and word_boundaries in page_text:
            pages_with_text.append(page_num)
            if not boundaries:
                words = plumber_reader.pages[page_num].extract_words()
                for word in words:
                    if word["text"] == word_boundaries:
                        boundaries = word["bottom"]

    pdf_document.close()
    return pages_with_text, boundaries

def proses_pdf_calk(pdf_path):
    pages_with_text, boundaries = search_pages(pdf_path, word_check)
    valid_calk = True
    with pdfplumber.open(pdf_path) as pdf:
        known_number_heading = []
        known_letter_heading = []
        foreign_key = ""
        
        if 'GENERAL' not in pdf.pages[pages_with_text[0]].extract_text():
            valid_calk = False
        print("\n[INFO] MELAKUKAN EKSTRAKSI HALAMAN")
        for page_num in pages_with_text:
            print(f"Mengekstrak Teks dari Halaman {page_num + 1}...")
            page = pdf.pages[page_num]
            
            if valid_calk:
                crop_box = (0, boundaries + 5, page.width - ((page.width//2) - 20), page.height - 30)
            else:
                crop_box = (0, boundaries + 5, page.width, page.height - 30)

            crop_page = page.within_bbox(crop_box)
            page_text = crop_page.extract_text()

            lines = page_text.split('\n')

            heading = ""
            content = ""
            next_line_heading = False
            next_letter_heading = False
            for line in lines:
                if next_line_heading and line.strip().isupper():
                    heading += " " + line.strip()
                    next_line_heading = False
                    continue
                elif next_letter_heading:
                    if line.strip()[0].islower():
                        heading += " " + line.strip()
                        continue
                    elif line.strip()[0].isupper():
                        next_letter_heading = False
                next_line_heading = False
                
                if any(line.strip().startswith(f"{i}.") for i in range(1, 100)):
                    case = line.strip().split('.')
                    if case[1].isupper() and len(case) == 2:
                        if heading and content:
                            data.append([foreign_key, heading, content.strip()])
                            content = ""
                        elif heading and not content:
                            data.append([foreign_key, heading, ""])
                        elif content:
                            if data:
                                data[-1][-1] = data[-1][-1] + " " + content.strip()
                            else:
                                data.append(["", "", content.strip()])
                            content = ""
                        
                        if case[0] not in known_number_heading:
                            heading = line.strip()
                            next_line_heading = True
                            known_number_heading.append(case[0])
                            known_letter_heading = []
                            foreign_key = case[0]
                elif re.compile(r"^[a-z]\.\s+").match(line.strip()):
                    case = line.strip().split('.')
                    valid = True
                    if case[0] not in known_letter_heading:
                        if known_letter_heading:
                            if case[0] != chr(ord(known_letter_heading[-1]) + 1):
                                valid = False
                                content += line.strip() + " "
                    if valid:
                        if heading and content:
                            data.append([foreign_key, heading, content.strip()])
                            content = ""
                        elif heading and not content:
                            data.append([foreign_key, heading, ""])
                        elif content:
                            if data:
                                data[-1][-1] = data[-1][-1] + " " + content.strip()
                            else:
                                data.append(["", "", content.strip()])
                            content = ""

                        if case[0] not in known_letter_heading:
                            heading = line.strip()
                            known_letter_heading.append(case[0])
                            next_letter_heading = True
                            if not foreign_key.isdigit():
                                foreign_key = foreign_key.replace(foreign_key[-1], case[0])
                            else:
                                foreign_key = foreign_key + case[0]
                        else:
                            heading = ""
                elif re.compile(r"^[a-z]{2}\.\s+").match(line.strip()):
                    case = line.strip().split('.')
                    valid = True
                    if case[0] not in known_letter_heading:
                        if "z" not in known_letter_heading:
                            valid = False
                            content += line.strip() + " "
                    if valid:
                        if heading and content:
                            data.append([foreign_key, heading, content.strip()])
                            content = ""
                        elif heading and not content:
                            data.append([foreign_key, heading, ""])
                        elif content:
                            if data:
                                data[-1][-1] = data[-1][-1] + " " + content.strip()
                            else:
                                data.append(["", "", content.strip()])
                            content = ""

                        if case[0] not in known_letter_heading:
                            heading = line.strip()
                            known_letter_heading.append(case[0])
                            next_letter_heading = True
                            if not foreign_key.isdigit():
                                foreign_key = foreign_key.replace(foreign_key[1:], case[0])
                            else:
                                foreign_key = foreign_key + case[0]
                        else:
                            heading = ""
                else:
                    content += line.strip() + " "

            if heading and content:
                data.append([foreign_key, heading, content.strip()])
            elif heading and not content:
                data.append([foreign_key, heading, ""])
            elif content:
                if data:
                    data[-1][-1] = data[-1][-1] + " " + content.strip()
                else:
                    data.append(["", "", content.strip()])
    masukkan_ke_database_calk()

def masukkan_ke_database_calk():
    # Cek Terlebih Dahulu Apakah CALK sudah masuk di Database
    if validasi_calk():
        messagebox.showwarning("Data Sudah Ada", "Data sudah ada sebelumnya!")
        return
    
    # Membuat Dataframe
    calk = pd.DataFrame(data, columns=["Kode", "Heading", "Content"])

    # Pengecekan jika ada content yang kosong (NULL) dan menyalin data content dari heading berikutnya
    for i in range(len(calk) - 1):
        if pd.isnull(calk.loc[i, "Content"]) or calk.loc[i, "Content"].strip() == "":
            # Jika content kosong, salin dari content baris berikutnya
            for j in range(i + 1, len(calk) - 1):
                if any(calk.loc[j, "Heading"].strip().startswith(f"{k}.") for k in range(1, 100)):
                    break
                calk.loc[i, "Content"] = calk.loc[i, "Content"] + " " + calk.loc[j, "Content"]

    # # Jika content pada baris terakhir kosong, ambil content dari heading terakhir
    if pd.isnull(calk.loc[len(calk) - 1, "Content"]) or calk.loc[len(calk) - 1, "Content"] == "":
        calk.loc[len(calk) - 1, "Content"] = calk.loc[len(calk) - 2, "Content"]

    try:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='db_keuangan'
        )

        # Hilangkan NaN
        calk = hilangkan_nan_strip(calk)

        for i in range(len(calk)):
            cursor = conn.cursor()
            sql = """
                INSERT INTO
                tb_catatan_atas_laporan_keuangan(kode_emiten, kode_calk, heading_calk, konten_calk)
                VALUES (%s, %s, %s, %s)
            """

            cursor.execute(sql, (kode_emiten_global, calk.iloc[i,0], calk.iloc[i,1], calk.iloc[i,2]))
            conn.commit()
        
        messagebox.showinfo("Berhasil", "Data Berhasil Disimpan: ")
        print("\n[INFO] MELAKUKAN PENYIMPANAN HASIL EKSTRAKSI HALAMAN")
        print("Data Berhasil Disimpan ke Dalam Database!")

        # Tutup Agar Aman
        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Error", f"Gagal Memasukkan ke Database: {e}")

# ======= Validasi UI ==========
def validasi_data_keuangan(grup_keuangan):
    global status_aruskas, status_neraca, status_labarugi

    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='db_keuangan'
    )

    cursor = conn.cursor()
    sql = """
        SELECT * 
        FROM tb_laporan_keuangan 
        WHERE 
        kode_emiten = %s AND tahun = %s AND quartal = %s AND grup_laporan_keuangan = %s
    """

    # Eksekusi query SELECT
    cursor.execute(sql, (informasi_tambahan[0], informasi_tambahan[2], informasi_tambahan[3], grup_keuangan))
    hasil_select = cursor.fetchone()  # Mengambil satu hasil query

    if hasil_select:
        if grup_keuangan == "Laporan Neraca":
            status_neraca = True
        elif grup_keuangan == "Laporan Laba Rugi":
            status_labarugi = True
        elif grup_keuangan == "Laporan Arus Kas":
            status_aruskas = True
        
        messagebox.showwarning("Data Sudah Ada", f"Data '{grup_keuangan}', pada {informasi_tambahan[0]} {informasi_tambahan[1]} {informasi_tambahan[2]} {informasi_tambahan[3]} Gagal Dimasukkan!")
    else:
        messagebox.showinfo("Informasi", f"Data '{grup_keuangan}', pada {informasi_tambahan[0]} {informasi_tambahan[1]} {informasi_tambahan[2]} {informasi_tambahan[3]} Berhasil Dimasukkan!")
    
    # Tutup Agar Aman
    cursor.close()
    conn.close()

def validasi_calk():
    try:
        conn = mysql.connector.connect(
            host='localhost',
            user='root',
            password='',
            database='db_keuangan'
        )

        cursor = conn.cursor()
        sql = """
            SELECT * 
            FROM tb_catatan_atas_laporan_keuangan
            WHERE 
            kode_emiten = %s
        """

        # Eksekusi query SELECT
        cursor.execute(sql, (kode_emiten_global,))
        hasil_select = cursor.fetchone()  # Mengambil satu hasil query

        if hasil_select:
            return True
        else:
            return False
        
        # Tutup Agar Aman
        cursor.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Error", f"Gagal Mengecek ke Database: {e}")

# ======= Kepentingan UI ==========
def hapus_file():
    # Dapatkan indeks item yang dipilih di Listbox
    selected_index = file_listbox.curselection()
    if not selected_index:
        messagebox.showwarning("Peringatan", "Pilih file yang ingin dihapus.")
        return

    # Dapatkan teks item yang dipilih
    selected_item = file_listbox.get(selected_index)

    # Hapus item dari Listbox
    file_listbox.delete(selected_index)

    # Hapus dari daftar selected_files_pdf atau selected_files_excel
    if selected_item.startswith("PDF: "):
        file_path = selected_item.replace("PDF: ", "")
        if file_path in selected_files_pdf:
            selected_files_pdf.remove(file_path)
    elif selected_item.startswith("Excel: "):
        file_path = selected_item.replace("Excel: ", "")
        if file_path in selected_files_excel:
            selected_files_excel.remove(file_path)

def tekan_proses():
    # Memproses File Excel dan PDF
    proses_files_excel()
    proses_files_pdf()

    # Menghapus seluruh path jika sudah diproses
    file_listbox.delete(0, tk.END)
    selected_files_pdf.clear()
    selected_files_excel.clear()
    informasi_tambahan.clear()
    informasi_sheet_keuangan.clear()

def tekan_calk():
    global kode_emiten_global
    # Mengambil File Path dan Memprosesnya Langsung
    for file_path in selected_files_pdf:
        try:
            temp_path = file_path

            # Hanya Abil nama dari path yang bersangkutan
            temp_path = Path(file_path).name

            kode_emiten_global = ambil_kode_emiten_nama_file(temp_path)
            proses_pdf_calk(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Gagal membaca file PDF: {e}")
    
    # Menghapus seluruh path jika sudah diproses
    file_listbox.delete(0, tk.END)
    selected_files_pdf.clear()
    selected_files_excel.clear()
    informasi_tambahan.clear()
    informasi_sheet_keuangan.clear()

# ============ Membuat UI menggunakan tkinter ==============
root = tk.Tk()
root.title("Mengambil Data Keuangan Bank")

# Tombol untuk menambahkan file PDF
tambah_file_pdf_button = tk.Button(root, text="Tambah File PDF", font=("Arial", 12), command=pilih_file_pdf)
tambah_file_pdf_button.pack(pady=10)

# Tombol untuk menambahkan file Excel
tambah_file_excel_button = tk.Button(root, text="Tambah File Excel", font=("Arial", 12), command=pilih_file_excel)
tambah_file_excel_button.pack(pady=10)

# Listbox untuk menampilkan daftar file yang dipilih
file_listbox = Listbox(root, width=160, height=20)
file_listbox.pack(pady=10)

# Tombol untuk menghapus file yang dipilih
hapus_file_button = tk.Button(root, text="Hapus File yang Dipilih", bg="red", fg="white", command=hapus_file)
hapus_file_button.pack(pady=10)

# Tombol untuk memproses file PDF CALK
proses_pdf_button = tk.Button(root, text="PROSES CALK!", bg="yellow", command=tekan_calk)
proses_pdf_button.pack(pady=10)

# Tombol untuk memproses file PDF
proses_pdf_button = tk.Button(root, text="PROSES FILE!", bg="green", font=("Arial", 12), fg="white", command=tekan_proses)
proses_pdf_button.pack(pady=10)

root.mainloop()